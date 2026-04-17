#!/usr/bin/env python3
"""
① 入社日列（H列）を削除
② 昇進履歴列（E列）の先頭に入社年月日を追記して上書き

完成フォーマット例:
  2016/5/1→入社 ／ 2019/6→係長 ／ 2022/4→工場長（課長） ／ 2024/6→工場長（次長）
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from pathlib import Path
from google_auth import get_sheets_service

SPREADSHEET_ID = '1qZbzvN8goWp1P_ORZuTxeQWnFQblE7B5Xq7k7V4RQY8'
SHEET_ID = 0
BASE_DIR = Path(__file__).parent

# ===================================================================
# 入社日データ
# ===================================================================

def load_products_join_dates() -> dict:
    """三幸プロダクツ名簿から {氏名: (年,月,日 or None)} を返す"""
    path = BASE_DIR / '総務・人事部' / '社員名簿（三幸プロダクツ）.xlsx'
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb['2024.04.01以降']
    result = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = row[3]
        join = row[7]
        if not name or not join:
            continue
        name_clean = str(name).replace('\u3000', '').strip()
        if hasattr(join, 'year'):
            result[name_clean] = (join.year, join.month, join.day)
        else:
            # "2022/10頃" のような文字列
            s = str(join).strip()
            parts = s.replace('頃', '').split('/')
            try:
                if len(parts) >= 2:
                    result[name_clean] = (int(parts[0]), int(parts[1]), None)
            except ValueError:
                pass
    return result


# 三幸商事 入社日（社員異動記録から特定済み）
SANKO_JOIN = {
    '深井智子':   (2012,  6, None),
    '片岡源太郎': (2014,  9, None),
    '秋月雅一':   (2015, 10, None),
    '柴田航':     (2017, 10, None),
    '大浦壮司':   (2018,  7, None),   # 有期スタート
    '倉元太郎':   (2019,  4,    1),
    '竹原黎':     (2019,  4,    1),
    '菅頌子':     (2019,  4,    1),
    '船吉紗矢':   (2019,  4,    1),
    '福田正志':   (2019,  5,   13),
    '田中マユミ': (2019,  5,   20),
    '池ヶ谷武弘': (2020,  4,    1),
    '飯塚理彩':   (2020,  4,    1),
    '鈴木祐介':   (2021,  5,   10),
    '大﨑紗佳':   (2022,  4,    1),
    '片岡隼人':   (2022,  4,    1),
    '角野百香':   (2023,  4,    1),
}

# ===================================================================
# 昇進履歴データ（add_promotion_history.py と同一）
# ===================================================================
PROMOTION_HISTORY = {
    '小澤泰明':   [(2013, 6, '工場長')],
    '新木巌人':   [(2013, 6, '係長'),   (2022, 6, '課長')],
    '豊國光成':   [(2013, 6, '係長'),   (2022, 6, '課長')],
    '畠平淳一':   [(2014, 6, '次長'),   (2017, 6, '部長'),  (2022, 8, '取締役（兼 部長）')],
    '竹村翔':     [(2014, 6, '係長'),   (2016, 6, '課長'),  (2019, 6, '次長'),
                  (2020, 6, '部長'),    (2022, 8, '取締役（兼 部長）')],
    '村上岳史':   [(2014, 6, '主任'),   (2019, 6, '係長'),  (2023, 6, '課長')],
    '池田浩':     [(2014, 6, '係長')],
    '内田剛人':   [(2014, 6, '主任')],
    '杉浦浩之':   [(2014, 6, '主任'),   (2021, 6, '係長')],
    '坂本秀勝':   [(2015, 6, '係長'),   (2019, 6, '課長')],
    '松本寛和':   [(2015, 6, '主任')],
    '小家和久':   [(2016, 6, '係長'),   (2022, 6, '課長')],
    '山口博志':   [(2016, 6, '係長')],
    '吉田雄亮':   [(2016, 6, '主任'),   (2021, 6, '係長')],
    '山崎正道':   [(2016, 6, '係長'),   (2020, 6, '課長')],
    '佐藤信吾':   [(2017, 6, '主任')],
    '徳本宏史':   [(2020, 6, '部長')],
    '伊藤大貴':   [(2020, 6, '係長')],
    '片岡源太郎': [(2021, 6, '係長')],
    '大野寛之':   [(2021, 6, '課長')],
    '秋月雅一':   [(2022, 6, '主任')],
    '木下和哉':   [(2022, 6, '主任')],
    '武田圭司':   [(2022, 6, '主任')],
    '若槻拓哉':   [(2024, 6, '係長')],
    '深井智子':   [(2024, 6, '主任')],
    '竹原黎':     [(2025, 6, '係長')],
    '船吉紗矢':   [(2025, 6, '主任')],
    '松谷匡晃':   [(2025, 6, '次長'),   (2025, 8, '東京支店長')],
    '吉田祐輔':   [(2019, 6, '係長'),   (2022, 4, '工場長（課長）'), (2024, 6, '工場長（次長）')],
    '齊藤三記夫': [(2021, 4, '工場長（課長）'), (2024, 6, '工場長（次長）')],
    '岡島優':     [(2022, 6, '主任')],
    '野澤哲央':   [(2023, 6, '主任')],
    '山田怜':     [(2024, 6, '主任')],
}


def fmt_join(year, month, day) -> str:
    """入社日エントリ文字列を生成"""
    if day:
        return f'{year}/{month}/{day}→入社'
    else:
        return f'{year}/{month}→入社'


def build_history(name: str, join_dates: dict) -> str:
    """入社日 + 昇進履歴を結合して文字列を返す"""
    # 入社日エントリ（sortキーとして (year, month, day or 0) を使用）
    entries = []   # [(sort_key, display_str), ...]

    join = join_dates.get(name)
    if join:
        y, m, d = join
        sort_key = (y, m, d or 0)
        entries.append((sort_key, fmt_join(y, m, d)))

    # 昇進エントリ
    for (y, m, role) in PROMOTION_HISTORY.get(name, []):
        entries.append(((y, m, 0), f'{y}/{m}→{role}'))

    if not entries:
        return ''

    entries.sort(key=lambda x: x[0])
    return ' ／ '.join(s for _, s in entries)


def main():
    # プロダクツ入社日を自動取得
    products_dates = load_products_join_dates()
    print(f'三幸プロダクツ 入社日取得: {len(products_dates)}名')

    # 全入社日マップ（三幸商事 + プロダクツ）
    all_join: dict = {}
    all_join.update(SANKO_JOIN)
    all_join.update(products_dates)

    service = get_sheets_service()

    # ---------- 1. 氏名列を取得 ----------
    result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range='シート1!A:A'
    ).execute()
    name_rows = result.get('values', [])

    # ---------- 2. 入社日列（H = index 7）を削除 ----------
    service.spreadsheets().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={
            'requests': [{
                'deleteDimension': {
                    'range': {
                        'sheetId': SHEET_ID,
                        'dimension': 'COLUMNS',
                        'startIndex': 7,   # H列（0-based）
                        'endIndex': 8,
                    }
                }
            }]
        }
    ).execute()
    print('H列（入社日）を削除しました')

    # ---------- 3. 昇進履歴列（E列）を入社日込みで上書き ----------
    data_updates = [
        {'range': 'シート1!E3', 'values': [['昇進履歴']]},
    ]

    written = 0
    for i, row in enumerate(name_rows):
        row_num = i + 1
        if row_num <= 3:
            continue
        if not row or not row[0].strip():
            continue
        name = row[0].strip()
        name_key = name.replace('\u3000', '').strip()

        # 入社日の名前照合（全角スペースあり・なし両方）
        join = all_join.get(name) or all_join.get(name_key)
        join_for_build = {name_key: join} if join else {}
        # build_historyに渡せるよう統合マップを作成
        lookup = {**all_join, name: join} if join else all_join

        hist = build_history(name_key if name_key in all_join or name_key in PROMOTION_HISTORY
                             else name, lookup)
        if hist:
            data_updates.append({
                'range': f'シート1!E{row_num}',
                'values': [[hist]]
            })
            written += 1
            print(f'  ✓ {name}: {hist}')
        else:
            # 既存の昇進履歴も入社日もない場合は空欄をセット（前の値を消す）
            data_updates.append({
                'range': f'シート1!E{row_num}',
                'values': [['']]
            })

    service.spreadsheets().values().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={
            'valueInputOption': 'USER_ENTERED',
            'data': data_updates
        }
    ).execute()

    print(f'\n=== 完了 ===')
    print(f'昇進履歴（入社日込み）を記入: {written} 名')


if __name__ == '__main__':
    main()
