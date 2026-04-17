#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
template_builder.py — openpyxl でテンプレートを白紙から構築するモジュール
三幸商事株式会社 経費精算書
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date as _date

FONT = "ＭＳ Ｐゴシック"
THIN = Side(style="thin")
HAIR = Side(style="hair")
DBLE = Side(style="double")
NONE = Side(style=None)


def _f(size=10, bold=False):
    return Font(name=FONT, size=size, bold=bold)


def _a(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _b(top=None, bot=None, left=None, right=None):
    return Border(
        top=top or NONE, bottom=bot or NONE,
        left=left or NONE, right=right or NONE
    )


def _colw(ws, col, w256):
    ws.column_dimensions[get_column_letter(col)].width = round(w256 / 256, 2)


def _rowh(ws, row, h20):
    ws.row_dimensions[row].height = round(h20 / 20, 2)


def _m(ws, cell_range):
    ws.merge_cells(cell_range)


def _box(ws, r1, c1, r2, c2, top=THIN, bot=THIN, left=THIN, right=THIN):
    """矩形領域の外枠に罫線を適用する"""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            at_t = (r == r1)
            at_b = (r == r2)
            at_l = (c == c1)
            at_r = (c == c2)
            if not (at_t or at_b or at_l or at_r):
                continue
            cell = ws.cell(row=r, column=c)
            eb = cell.border
            cell.border = Border(
                top=top if at_t else eb.top,
                bottom=bot if at_b else eb.bottom,
                left=left if at_l else eb.left,
                right=right if at_r else eb.right,
            )


def _label(ws, cell_ref, text, size=10, bold=False,
           h="center", v="center", wrap=False):
    c = ws[cell_ref]
    c.value = text
    c.font = _f(size, bold)
    c.alignment = _a(h, v, wrap)
    return c


# ============================================================
# 交際費（接待交際費申請書）
# ============================================================
def build_kosaihi(data):
    """
    data keys:
        申請日: date, 申請者: str,
        相手先会社: str, 相手先出席者: str, 相手先人数: int,
        関係: str, 自社出席者: str, 自社人数: int,
        実施日: date, 実施時間: str,
        場所名: str, 場所住所: str, 目的: str,
        expenses: [(支払先, 内訳, 実績額), ...] (max 5),
        仮払金: int
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "接待交際費申請書"
    ws.page_setup.paperSize = 9   # A4
    ws.page_setup.orientation = "portrait"
    ws.print_area = "A1:AD44"

    # 列幅 (576=2.25char, Z列のみ512=2.0char)
    for col in range(1, 31):
        _colw(ws, col, 512 if col == 26 else 576)

    # 行高
    for row in range(1, 45):
        if row == 7:      h20 = 300
        elif row == 8:    h20 = 162
        elif 29 <= row <= 38: h20 = 319
        else:             h20 = 259
        _rowh(ws, row, h20)

    # タイトル
    _m(ws, "A1:AD2")
    _label(ws, "A1", "接待交際費申請書", size=14)

    # 印鑑欄 行3 (ラベル) + 行4-7 (空欄)
    for rng, lbl in [("A3:E3", "社長"), ("F3:J3", "所属長"), ("K3:O3", "検印")]:
        _m(ws, rng)
        r, c = rng.split(":")[0], rng.split(":")[0]
        _label(ws, c, lbl)
    _box(ws, 3, 1, 3, 5);  _box(ws, 3, 6, 3, 10);  _box(ws, 3, 11, 3, 15)
    _m(ws, "A4:E7");  _m(ws, "F4:J7");  _m(ws, "K4:O7")
    _box(ws, 4, 1, 7, 5);  _box(ws, 4, 6, 7, 10);  _box(ws, 4, 11, 7, 15)

    # 申請日/申請者
    _m(ws, "R4:U4");  _label(ws, "R4", "申請日")
    _m(ws, "V4:AD4")
    ws["V4"].value = _date.today()   # 申請日は作成日を自動セット
    ws["V4"].number_format = "YYYY/MM/DD"
    ws["V4"].font = _f(10);  ws["V4"].alignment = _a()

    _m(ws, "R6:U7")
    _label(ws, "R6", "申請者\n氏　　名", wrap=True)
    _m(ws, "V6:AC7")
    _label(ws, "V6", data.get("申請者", ""), size=11)

    # 行7 R-AD 下線
    for c in range(18, 31):
        cell = ws.cell(row=7, column=c)
        eb = cell.border
        cell.border = Border(top=eb.top, bottom=THIN, left=eb.left, right=eb.right)

    # 相手先名及び出席者氏名 行9-13
    _m(ws, "A9:E13")
    _label(ws, "A9", "相手先名及び\n出席者氏名", size=9, h="distributed", wrap=True)
    _box(ws, 9, 1, 13, 5)

    _m(ws, "F9:H10");  _label(ws, "F9", "会社名");  _box(ws, 9, 6, 10, 8)
    _m(ws, "I9:AA10")
    ws["I9"].value = data.get("相手先会社", "")
    ws["I9"].font = _f(); ws["I9"].alignment = _a("general", "center")
    _box(ws, 9, 9, 10, 27)

    _m(ws, "AB9:AD9");  _label(ws, "AB9", "人数");  _box(ws, 9, 28, 9, 30)
    _m(ws, "AB10:AD13")
    ws["AB10"].value = data.get("相手先人数", "")
    ws["AB10"].font = _f(); ws["AB10"].alignment = _a()
    _box(ws, 10, 28, 13, 30)

    _m(ws, "F11:H13");  _label(ws, "F11", "出席者");  _box(ws, 11, 6, 13, 8)
    _m(ws, "I11:AA13")
    ws["I11"].value = data.get("相手先出席者", "")
    ws["I11"].font = _f(); ws["I11"].alignment = _a("general", "center")
    _box(ws, 11, 9, 13, 27)

    # 当社との関係 行14-15
    _m(ws, "A14:H15")
    _label(ws, "A14", "当社との関係", h="distributed", v="distributed")
    _box(ws, 14, 1, 15, 8)
    _m(ws, "I14:AD15")
    ws["I14"].value = data.get("関係", "得意先")
    ws["I14"].font = _f(); ws["I14"].alignment = _a("general", "center")
    _box(ws, 14, 9, 15, 30)

    # 当社出席者 行16-18
    _m(ws, "AB16:AD16");  _label(ws, "AB16", "人数");  _box(ws, 16, 28, 16, 30)
    _m(ws, "A16:H18")
    _label(ws, "A16", "当社出席者", h="distributed", v="distributed")
    _box(ws, 16, 1, 18, 8)
    _m(ws, "I16:AA18")
    ws["I16"].value = data.get("自社出席者", "")
    ws["I16"].font = _f(); ws["I16"].alignment = _a("general", "center")
    _box(ws, 16, 9, 18, 27)
    _m(ws, "AB17:AD20")
    ws["AB17"].value = data.get("自社人数", "")
    ws["AB17"].font = _f(); ws["AB17"].alignment = _a()
    _box(ws, 17, 28, 20, 30)

    # 実施日時 行19-20
    _m(ws, "A19:H20")
    _label(ws, "A19", "実施日時", h="distributed", v="distributed")
    _box(ws, 19, 1, 20, 8)
    _m(ws, "I19:T20")
    if data.get("実施日"):
        ws["I19"].value = data["実施日"]
        ws["I19"].number_format = "YYYY/MM/DD"
    ws["I19"].font = _f(); ws["I19"].alignment = _a()
    _box(ws, 19, 9, 20, 20)
    _m(ws, "U19:U20");  _label(ws, "U19", "時間", size=9);  _box(ws, 19, 21, 20, 21)
    _m(ws, "V19:AA20")
    ws["V19"].value = data.get("実施時間", "")
    ws["V19"].font = _f(); ws["V19"].alignment = _a()
    _box(ws, 19, 22, 20, 27)

    # 接待場所 行21-24
    _m(ws, "AB21:AD21");  _label(ws, "AB21", "人数計");  _box(ws, 21, 28, 21, 30)
    _m(ws, "A21:E24")
    _label(ws, "A21", "接待場所", h="distributed", v="distributed")
    _box(ws, 21, 1, 24, 5)
    _m(ws, "F21:H22")
    _label(ws, "F21", "名称", v="distributed"); _box(ws, 21, 6, 22, 8)
    _m(ws, "I21:AA22")
    ws["I21"].value = data.get("場所名", "")
    ws["I21"].font = _f(); ws["I21"].alignment = _a("general", "center")
    _box(ws, 21, 9, 22, 27)
    _m(ws, "AB22:AD24")
    ws["AB22"].value = (data.get("相手先人数") or 0) + (data.get("自社人数") or 0)
    ws["AB22"].font = _f(bold=True); ws["AB22"].alignment = _a()
    _box(ws, 22, 28, 24, 30)
    _m(ws, "F23:H24")
    _label(ws, "F23", "住所", v="distributed"); _box(ws, 23, 6, 24, 8)
    _m(ws, "I23:AA24")
    ws["I23"].value = data.get("場所住所", "")
    ws["I23"].font = _f(); ws["I23"].alignment = _a("general", "center", wrap=True)
    _box(ws, 23, 9, 24, 27)

    # 目的 行25-27
    _m(ws, "A25:H27")
    _label(ws, "A25", "目的", h="distributed", v="distributed")
    _box(ws, 25, 1, 27, 8)
    _m(ws, "I25:AD27")
    ws["I25"].value = data.get("目的", "")
    ws["I25"].font = _f(); ws["I25"].alignment = _a("general", "center", wrap=True)
    _box(ws, 25, 9, 27, 30)

    # 支出の内容 行28 (ヘッダーラベル、罫線なし)
    ws.cell(row=28, column=1).value = "支出の内容"
    ws.cell(row=28, column=1).font = _f()
    ws.cell(row=28, column=1).alignment = _a("general", "center")

    # 明細テーブル ヘッダー 行29
    for rng, lbl, c1, c2 in [
        ("A29:G29",  "支払先", 1,  7),
        ("H29:O29",  "内訳",   8,  15),
        ("P29:T29",  "予算額", 16, 20),
        ("U29:Y29",  "実績額", 21, 25),
        ("Z29:AD29", "差異",   26, 30),
    ]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl)
        _box(ws, 29, c1, 29, c2)

    # 明細行 行30-34 (最大5行)
    expenses = data.get("expenses", []) or []
    total = 0
    for i in range(5):
        r = 30 + i
        for rng, c1, c2 in [
            (f"A{r}:G{r}", 1, 7), (f"H{r}:O{r}", 8, 15),
            (f"P{r}:T{r}", 16, 20), (f"U{r}:Y{r}", 21, 25),
            (f"Z{r}:AD{r}", 26, 30),
        ]:
            _m(ws, rng)
            ws[rng.split(":")[0]].font = _f()
            ws[rng.split(":")[0]].alignment = _a("general", "center")
        top_s = THIN if i == 0 else HAIR
        bot_s = THIN if i == 4 else HAIR
        for c1, c2 in [(1,7),(8,15),(16,20),(21,25),(26,30)]:
            _box(ws, r, c1, r, c2, top=top_s, bot=bot_s)
        if i < len(expenses):
            payee, detail, amount = expenses[i]
            amount = amount or 0
            ws[f"A{r}"].value = payee
            ws[f"H{r}"].value = detail
            if amount:
                ws[f"U{r}"].value = amount
            ws[f"Z{r}"].value = 0
            total += amount

    # 支払合計 行35
    _m(ws, "A35:O35");  _label(ws, "A35", "支払合計");  _box(ws, 35, 1, 35, 15)
    _m(ws, "P35:T35");  ws["P35"].value = 0;  ws["P35"].font = _f()
    ws["P35"].alignment = _a("general", "center");  _box(ws, 35, 16, 35, 20)
    _m(ws, "U35:Y35")
    ws["U35"].value = total; ws["U35"].font = _f(bold=True)
    ws["U35"].alignment = _a("general", "center");  _box(ws, 35, 21, 35, 25)
    _m(ws, "Z35:AD35");  ws["Z35"].value = 0;  ws["Z35"].font = _f()
    ws["Z35"].alignment = _a("general", "center");  _box(ws, 35, 26, 35, 30)

    # 仮払金 行36
    karibara = data.get("仮払金") or 0
    _m(ws, "A36:T36");  _label(ws, "A36", "仮払金");  _box(ws, 36, 1, 36, 20)
    _m(ws, "U36:Y36")
    if karibara: ws["U36"].value = karibara
    ws["U36"].font = _f(); ws["U36"].alignment = _a("general", "center")
    _box(ws, 36, 21, 36, 25)
    _m(ws, "Z36:AD36")
    _label(ws, "Z36", "1人当たり飲食費", size=8);  _box(ws, 36, 26, 36, 30)

    # 精算額 行37
    _m(ws, "A37:T37");  _label(ws, "A37", "精算額");  _box(ws, 37, 1, 37, 20)
    _m(ws, "U37:Y37")
    ws["U37"].value = total - karibara; ws["U37"].font = _f(bold=True)
    ws["U37"].alignment = _a("general", "center");  _box(ws, 37, 21, 37, 25)
    _m(ws, "Z37:AD37");  ws["Z37"].value = 0;  ws["Z37"].font = _f(bold=True)
    ws["Z37"].alignment = _a("general", "center");  _box(ws, 37, 26, 37, 30)

    # （経理課） 行38
    ws.cell(row=38, column=2).value = "（経理課）"
    ws.cell(row=38, column=2).font = _f()
    ws.cell(row=38, column=2).alignment = _a("general", "center")

    # 経理責任者/検印/出納 行39
    for rng, lbl, c1, c2 in [
        ("A39:E39", "経理責任者", 1, 5),
        ("F39:J39", "検印",       6, 10),
        ("K39:O39", "出納",       11, 15),
    ]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl); _box(ws, 39, c1, 39, c2)

    # 領収日 行39 R-AD (下線のみ) — 作成日を自動セット
    _m(ws, "R39:U39");  _label(ws, "R39", "領収日", h="general")
    _m(ws, "V39:AD39")
    ws["V39"].value = _date.today()
    ws["V39"].number_format = "YYYY/MM/DD"
    ws["V39"].font = _f(); ws["V39"].alignment = _a()
    for c in range(18, 31):
        cell = ws.cell(row=39, column=c)
        eb = cell.border
        cell.border = Border(top=eb.top, bottom=THIN, left=eb.left, right=eb.right)

    # 印/申請者名 行40-41
    _m(ws, "R40:V41");  _m(ws, "W40:AD41")
    申請者 = data.get("申請者", "")
    ws["W40"].value = f"{申請者}　　　　印"
    ws["W40"].font = _f(9); ws["W40"].alignment = _a("right", "center")

    # 下部印鑑欄 行40-43
    _m(ws, "A40:E43");  _m(ws, "F40:J43");  _m(ws, "K40:O43")
    _box(ws, 40, 1, 43, 5);  _box(ws, 40, 6, 43, 10);  _box(ws, 40, 11, 43, 15)

    # 上記金額 行42
    ws.cell(row=42, column=17).value = "上記金額正に領収致しました。"
    ws.cell(row=42, column=17).font = _f()
    ws.cell(row=42, column=17).alignment = _a("general", "center")

    return wb


# ============================================================
# 旅費交通費（旅費交通費精算書）
# ============================================================
def build_kotsu(data):
    """
    data keys:
        精算日: date, 報告者: str,
        目的地: str, 目的: str,
        routes: [(date, 行き先, 支払先, 内訳, 金額), ...] (max 12),
        日当フラグ: bool, 出張開始: date, 出張終了: date, 日数: int, 日当単価: int,
        宿泊先: str, 宿泊料: int,
        misc: [(date, 支払先, 摘要, 金額), ...] (max 5),
        仮払金: int
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "旅費交通費精算書"
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.print_area = "A1:AE46"

    # 列幅 (A-D=576, E-K=704, L-R=576, S=672, T-AE=576)
    for col in range(1, 32):
        if 5 <= col <= 11:   w = 704
        elif col == 19:       w = 672  # S
        else:                 w = 576
        _colw(ws, col, w)

    # 行高
    ROW_H = {
        1: 222, 2: 222, 3: 282, 4: 282, 5: 240, 6: 240,
        7: 255, 8: 222, 9: 255, 10: 240, 11: 240, 12: 240,
        13: 319, 40: 120, 41: 319,
    }
    for row in range(1, 47):
        if row in ROW_H:
            h20 = ROW_H[row]
        elif 14 <= row <= 25:
            h20 = 465
        elif 26 <= row <= 39:
            h20 = 319
        else:
            h20 = 259
        _rowh(ws, row, h20)

    # タイトル
    _m(ws, "A1:AD2");  _label(ws, "A1", "旅費交通費精算書", size=14)

    # 印鑑欄 行3-7
    for rng, lbl in [("A3:E3", "所属長"), ("F3:J3", "検印"), ("K3:O3", "検印")]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl, size=9)
    _box(ws, 3, 1, 3, 5);  _box(ws, 3, 6, 3, 10);  _box(ws, 3, 11, 3, 15)
    _m(ws, "A4:E7");  _m(ws, "F4:J7");  _m(ws, "K4:O7")
    _box(ws, 4, 1, 7, 5);  _box(ws, 4, 6, 7, 10);  _box(ws, 4, 11, 7, 15)

    # 精算日
    _m(ws, "Q3:S3");  _label(ws, "Q3", "精算日", v="center")
    # 下線（精算日値）
    _m(ws, "T3:AD3")
    if data.get("精算日"):
        ws["T3"].value = data["精算日"]; ws["T3"].number_format = "YYYY/MM/DD"
    ws["T3"].font = _f(); ws["T3"].alignment = _a()
    for c in range(17, 31):
        cell = ws.cell(row=3, column=c)
        eb = cell.border
        cell.border = Border(top=eb.top, bottom=THIN, left=eb.left, right=eb.right)

    # 報告者
    _m(ws, "Q4:S4");  _label(ws, "Q4", "報告者")
    _m(ws, "T4:AD4")
    ws["T4"].value = data.get("報告者", ""); ws["T4"].font = _f()
    ws["T4"].alignment = _a()
    _m(ws, "Q5:S6");  _label(ws, "Q5", "報告者")  # duplicate row 5-6 label
    _m(ws, "T5:AB6")
    ws["T5"].value = data.get("報告者", ""); ws["T5"].font = _f()
    ws["T5"].alignment = _a()
    _m(ws, "AC5:AD6")
    _label(ws, "AC5", "印", size=9)

    # 目的地 行8-9
    _m(ws, "A8:D9")
    _label(ws, "A8", "目的地", v="center")
    ws["A8"].border = _b(top=DBLE, left=DBLE)
    _m(ws, "E8:AD9")
    ws["E8"].value = data.get("目的地", "")
    ws["E8"].font = _f(9); ws["E8"].alignment = _a("general", "center", wrap=True)
    ws["E8"].border = _b(top=DBLE, left=THIN)
    ws.cell(row=9, column=1).border = _b(left=DBLE)
    ws.cell(row=9, column=4).border = _b(right=THIN)

    # 目的 行10-12
    _m(ws, "A10:D12");  _label(ws, "A10", "目的")
    ws["A10"].border = _b(top=THIN, left=DBLE)
    _m(ws, "E10:AD12")
    ws["E10"].value = data.get("目的", "")
    ws["E10"].font = _f(); ws["E10"].alignment = _a("general", "center", wrap=True)
    ws["E10"].border = _b(top=THIN, left=THIN)
    ws.cell(row=10, column=1).border = _b(top=THIN, left=DBLE)
    ws.cell(row=12, column=1).border = _b(left=DBLE)
    ws.cell(row=12, column=4).border = _b(right=THIN)

    # 交通費ヘッダー 行13
    for rng, lbl, c1, c2 in [
        ("A13:D13", "月日",   1,  4),
        ("E13:K13", "行き先", 5,  11),
        ("L13:R13", "支払先", 12, 18),
        ("S13:Z13", "内訳",   19, 26),
        ("AA13:AD13", "金　額", 27, 30),
    ]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl)
        _box(ws, 13, c1, 13, c2, left=DBLE if c1 == 1 else THIN)

    # 交通費明細 行14-25 (最大12行)
    routes = data.get("routes", []) or []
    total_fare = 0
    for i in range(12):
        r = 14 + i
        for rng in [f"A{r}:D{r}", f"E{r}:K{r}", f"L{r}:R{r}", f"S{r}:Z{r}", f"AA{r}:AD{r}"]:
            _m(ws, rng)
        top_s = HAIR if i > 0 else THIN
        _box(ws, r, 1, r, 4, top=top_s, bot=HAIR, left=DBLE)
        for c1, c2 in [(5, 11), (12, 18), (19, 26), (27, 30)]:
            _box(ws, r, c1, r, c2, top=top_s, bot=HAIR)
        if i < len(routes):
            d, route, payer, detail, amount = routes[i]
            amount = amount or 0
            ws[f"A{r}"].value = d.strftime("%m/%d") if d else ""
            ws[f"E{r}"].value = route
            ws[f"L{r}"].value = payer
            ws[f"S{r}"].value = detail
            if amount: ws[f"AA{r}"].value = amount
            ws[f"AA{r}"].alignment = _a("right", "center")
            total_fare += amount
        for ref in [f"A{r}", f"E{r}", f"L{r}", f"S{r}", f"AA{r}"]:
            c = ws[ref]; c.font = _f(8 if ref.startswith(("E", "L", "S")) else 10)
            if not c.alignment.horizontal: c.alignment = _a("general", "center", wrap=True)

    # 交通費合計 行26
    _m(ws, "A26:Z26");  _label(ws, "A26", "交通費合計①", bold=True)
    ws["A26"].border = _b(top=THIN, left=DBLE, right=THIN)
    _m(ws, "AA26:AD26")
    ws["AA26"].value = total_fare; ws["AA26"].font = _f(bold=True)
    ws["AA26"].alignment = _a("right", "center"); _box(ws, 26, 27, 26, 30, bot=HAIR)

    # 日当 行27
    _m(ws, "A27:C27");  _label(ws, "A27", "日　当")
    ws["A27"].border = _b(top=THIN, bot=THIN, left=DBLE, right=THIN)
    _m(ws, "D27:G27");  _label(ws, "D27", "出張期間");  _box(ws, 27, 4, 27, 7)
    _m(ws, "H27:K27")
    if data.get("出張開始"):
        ws["H27"].value = data["出張開始"]; ws["H27"].number_format = "YYYY/MM/DD"
    ws["H27"].font = _f(9); ws["H27"].alignment = _a(); _box(ws, 27, 8, 27, 11)
    ws.cell(row=27, column=12).value = "～"
    ws.cell(row=27, column=12).font = _f(9); ws.cell(row=27, column=12).alignment = _a()
    _box(ws, 27, 12, 27, 12)
    _m(ws, "M27:P27")
    if data.get("出張終了"):
        ws["M27"].value = data["出張終了"]; ws["M27"].number_format = "YYYY/MM/DD"
    ws["M27"].font = _f(9); ws["M27"].alignment = _a(); _box(ws, 27, 13, 27, 16)
    _m(ws, "Q27:R27");  _label(ws, "Q27", "日数");  _box(ws, 27, 17, 27, 18)
    _m(ws, "S27:T27")
    ws["S27"].value = data.get("日数") or 0; ws["S27"].font = _f(); ws["S27"].alignment = _a()
    _box(ws, 27, 19, 27, 20)
    _m(ws, "U27:W27");  _label(ws, "U27", "単価");  _box(ws, 27, 21, 27, 23)
    _m(ws, "X27:Z27")
    ws["X27"].value = data.get("日当単価") or 5000; ws["X27"].font = _f(); ws["X27"].alignment = _a()
    _box(ws, 27, 24, 27, 26)
    _m(ws, "AA27:AD27")
    nittou = (data.get("日数") or 0) * (data.get("日当単価") or 5000) if data.get("日当フラグ") else 0
    if nittou: ws["AA27"].value = nittou
    ws["AA27"].font = _f(); ws["AA27"].alignment = _a("general", "center"); _box(ws, 27, 27, 27, 30, bot=HAIR)

    # 宿泊料 行28-29
    _m(ws, "A28:C29");  _label(ws, "A28", "宿泊料")
    ws["A28"].border = _b(bot=HAIR, left=DBLE, right=THIN)
    _m(ws, "D28:G29");  _label(ws, "D28", "宿泊先");  _box(ws, 28, 4, 29, 7, bot=HAIR)
    _m(ws, "H28:Z28");  _m(ws, "H29:Z29")
    ws["H28"].value = data.get("宿泊先") or ""
    ws["H28"].font = _f(); ws["H28"].alignment = _a("general", "center")
    _box(ws, 28, 8, 29, 26, bot=HAIR)
    _m(ws, "AA28:AD28");  _m(ws, "AA29:AD29")
    if data.get("宿泊料"): ws["AA28"].value = data["宿泊料"]
    ws["AA28"].font = _f(); ws["AA28"].alignment = _a("general", "center"); _box(ws, 28, 27, 29, 30, bot=HAIR)
    ws.cell(row=29, column=1).border = _b(left=DBLE)
    ws.cell(row=29, column=4).border = _b(left=THIN)

    # 旅費交通費合計② 行30
    total_travel = total_fare + nittou + (data.get("宿泊料") or 0)
    _m(ws, "A30:Z30");  _label(ws, "A30", "旅費交通費合計②", bold=True)
    ws["A30"].border = _b(bot=THIN, left=DBLE, right=THIN)
    _m(ws, "AA30:AD30")
    ws["AA30"].value = total_travel; ws["AA30"].font = _f(bold=True)
    ws["AA30"].alignment = _a("general", "center"); _box(ws, 30, 27, 30, 30)

    # 諸経費ヘッダー 行31
    _m(ws, "A31:B36")  # 左ラベル（縦結合）
    _label(ws, "A31", "諸経費"); ws["A31"].border = _b(top=THIN, bot=HAIR, left=DBLE, right=THIN)
    _m(ws, "C31:F31");  _label(ws, "C31", "月日");  _box(ws, 31, 3, 31, 6)
    _m(ws, "G31:O31");  _label(ws, "G31", "支払先");  _box(ws, 31, 7, 31, 15)
    _m(ws, "P31:Z31");  _label(ws, "P31", "摘　要");  _box(ws, 31, 16, 31, 26)
    _m(ws, "AA31:AD31"); _label(ws, "AA31", "金　額"); _box(ws, 31, 27, 31, 30)

    # 諸経費明細 行32-35
    misc = data.get("misc", []) or []
    total_misc = 0
    for i in range(4):
        r = 32 + i
        _m(ws, f"C{r}:F{r}"); _m(ws, f"G{r}:O{r}"); _m(ws, f"P{r}:Z{r}"); _m(ws, f"AA{r}:AD{r}")
        top_s = HAIR
        bot_s = THIN if i == 3 else HAIR
        ws.cell(row=r, column=1).border = _b(bot=bot_s, left=DBLE)
        ws.cell(row=r, column=2).border = _b(bot=bot_s)
        ws.cell(row=r, column=3).border = _b(top=top_s, bot=bot_s, left=THIN, right=THIN)
        ws.cell(row=r, column=7).border = _b(top=top_s, bot=bot_s, left=THIN, right=THIN)
        ws.cell(row=r, column=16).border = _b(top=top_s, bot=bot_s, left=THIN, right=THIN)
        _box(ws, r, 27, r, 30, top=top_s, bot=bot_s)
        for ref in [f"C{r}", f"G{r}", f"P{r}", f"AA{r}"]:
            ws[ref].font = _f(); ws[ref].alignment = _a("general", "center")
        if i < len(misc):
            md, mpayee, mdesc, mamount = misc[i]
            mamount = mamount or 0
            ws[f"C{r}"].value = md.strftime("%m/%d") if md else ""
            ws[f"G{r}"].value = mpayee
            ws[f"P{r}"].value = mdesc
            if mamount: ws[f"AA{r}"].value = mamount
            total_misc += mamount

    # 諸経費合計③ 行36
    _m(ws, "C36:Z36");  _label(ws, "C36", "諸経費合計③", bold=True)
    ws.cell(row=36, column=1).border = _b(bot=THIN, left=DBLE)
    ws.cell(row=36, column=2).border = _b(bot=THIN)
    ws["C36"].border = _b(bot=THIN, left=THIN, right=THIN)
    _m(ws, "AA36:AD36")
    ws["AA36"].value = total_misc; ws["AA36"].font = _f(bold=True)
    ws["AA36"].alignment = _a("general", "center"); _box(ws, 36, 27, 36, 30)

    # 総合計④ 行37
    _m(ws, "A37:Z37");  _label(ws, "A37", "総合計④（②＋③）", bold=True)
    ws["A37"].border = _b(bot=HAIR, left=DBLE, right=THIN)
    _m(ws, "AA37:AD37")
    ws["AA37"].value = total_travel + total_misc; ws["AA37"].font = _f(bold=True)
    ws["AA37"].alignment = _a("general", "center"); _box(ws, 37, 27, 37, 30, bot=HAIR)

    # 仮払金 行38
    karibara = data.get("仮払金") or 0
    _m(ws, "A38:Z38");  _label(ws, "A38", "仮払金")
    ws["A38"].border = _b(top=HAIR, bot=THIN, left=DBLE, right=THIN)
    _m(ws, "AA38:AD38")
    if karibara: ws["AA38"].value = karibara
    ws["AA38"].font = _f(); ws["AA38"].alignment = _a("general", "center"); _box(ws, 38, 27, 38, 30)

    # 差引精算額 行39
    _m(ws, "A39:Z39");  _label(ws, "A39", "差引精算額", bold=True)
    ws["A39"].border = _b(bot=DBLE, left=DBLE, right=THIN)
    _m(ws, "AA39:AD39")
    ws["AA39"].value = total_travel + total_misc - karibara; ws["AA39"].font = _f(bold=True)
    ws["AA39"].alignment = _a("general", "center"); _box(ws, 39, 27, 39, 30, bot=DBLE)

    # 行40 スペーサー
    ws.cell(row=40, column=2).value = "（経理課）"
    ws.cell(row=40, column=2).font = _f(); ws.cell(row=40, column=2).alignment = _a("general", "center")

    # 下部印鑑欄 行41-45
    _m(ws, "E40:AD40")
    for rng, lbl, c1, c2 in [
        ("F41:J41", "所属長", 6, 10),
        ("K41:O41", "検印",   11, 15),
        ("P41:T41", "出納",   16, 20),
    ]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl); _box(ws, 41, c1, 41, c2)
    for rng, c1, c2 in [("F42:J45", 6, 10), ("K42:O45", 11, 15), ("P42:T45", 16, 20)]:
        _m(ws, rng); _box(ws, 42, c1, 45, c2)

    return wb


# ============================================================
# 金銭受領書（金銭請求書兼領収書）
# ============================================================
def build_kinsen(data):
    """
    data keys:
        届出日: date, 氏名: str, 目的: str,
        items: [(date, 支払先, 摘要, 金額), ...] (max 6),
        仮払金: int
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "金銭請求書兼領収書"
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.print_area = "A1:AC33"

    # 列幅 (全列 576=2.25)
    for col in range(1, 30):
        _colw(ws, col, 576)

    # 行高
    ROW_H = {
        1: 240, 2: 240, 3: 199, 4: 300, 5: 270, 6: 240,
        7: 240, 8: 270, 9: 240, 10: 240, 11: 240,
        12: 240, 13: 240, 14: 240, 15: 240, 16: 240,
        17: 402,  # ヘッダー行
        24: 402, 25: 402, 26: 402,  # 合計行
        27: 285, 28: 300,
    }
    for row in range(1, 34):
        if row in ROW_H:
            h20 = ROW_H[row]
        elif 18 <= row <= 23:
            h20 = 600   # 明細行
        else:
            h20 = 240
        _rowh(ws, row, h20)

    # タイトル
    _m(ws, "A1:AC2");  _label(ws, "A1", "金銭請求書兼領収書", size=12)

    # 届出日/氏名 行3-8
    _m(ws, "Q3:T4");  _label(ws, "Q3", "届出日")
    _m(ws, "U3:AC4")
    if data.get("届出日"):
        ws["U3"].value = data["届出日"]; ws["U3"].number_format = "YYYY/MM/DD"
    ws["U3"].font = _f(); ws["U3"].alignment = _a()

    _m(ws, "A4:E4");  _m(ws, "F4:J4");  _m(ws, "K4:O4")
    for lbl, rng, c1, c2 in [("所属長","A4:E4",1,5), ("検印","F4:J4",6,10), ("検印","K4:O4",11,15)]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl); _box(ws, 4, c1, 4, c2)
    _m(ws, "A5:E8");  _m(ws, "F5:J8");  _m(ws, "K5:O8")
    _box(ws, 5, 1, 8, 5);  _box(ws, 5, 6, 8, 10);  _box(ws, 5, 11, 8, 15)

    _m(ws, "Q5:T6");  _label(ws, "Q5", "氏　名")
    _m(ws, "U5:AC6")
    ws["U5"].value = data.get("氏名", ""); ws["U5"].font = _f(); ws["U5"].alignment = _a()
    _m(ws, "Q8:AC8")
    _label(ws, "Q8", "下記の金額の出金をお願い致します。")

    # 金額 行10-11
    _m(ws, "C10:I11");  _label(ws, "C10", "金　　額", size=14, bold=True)
    _box(ws, 10, 3, 11, 9)
    _m(ws, "J10:AB11")
    items = data.get("items", []) or []
    total = sum((it[3] or 0) for it in items)
    ws["J10"].value = total; ws["J10"].font = _f(14, bold=True)
    ws["J10"].alignment = _a()
    ws["J10"].number_format = '#,##0"円"'
    _box(ws, 10, 10, 11, 28)

    # 使用目的 行13-16
    _m(ws, "A13:D16");  _label(ws, "A13", "使用目的")
    ws["A13"].border = _b(top=DBLE, left=DBLE)
    _m(ws, "E13:AC16")
    ws["E13"].value = data.get("目的", "")
    ws["E13"].font = _f(); ws["E13"].alignment = _a("general", "top", wrap=True)
    ws["E13"].border = _b(top=DBLE, left=THIN)
    for r in range(13, 17):
        ws.cell(row=r, column=1).border = _b(top=DBLE if r==13 else None, left=DBLE)
        ws.cell(row=r, column=4).border = _b(top=DBLE if r==13 else None, right=THIN)

    # 明細ヘッダー 行17
    for rng, lbl, c1, c2 in [
        ("A17:D17",  "月日",   1,  4),
        ("E17:K17",  "支払先", 5,  11),
        ("L17:X17",  "摘　要", 12, 24),
        ("Y17:AC17", "金額",   25, 29),
    ]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl)
        _box(ws, 17, c1, 17, c2, left=DBLE if c1 == 1 else THIN)

    # 明細行 行18-23 (最大6行)
    total_misc = 0
    for i in range(6):
        r = 18 + i
        for rng in [f"A{r}:D{r}", f"E{r}:K{r}", f"L{r}:X{r}", f"Y{r}:AC{r}"]:
            _m(ws, rng)
        top_s = NONE if i == 0 else HAIR
        bot_s = HAIR
        _box(ws, r, 1, r, 4, top=top_s, bot=bot_s, left=DBLE)
        for c1, c2 in [(5, 11), (12, 24), (25, 29)]:
            _box(ws, r, c1, r, c2, top=top_s, bot=bot_s)
        for ref in [f"A{r}", f"E{r}", f"L{r}", f"Y{r}"]:
            ws[ref].font = _f(); ws[ref].alignment = _a("general", "center", wrap=True)
        if i < len(items):
            d, payee, desc, amount = items[i]
            amount = amount or 0
            ws[f"A{r}"].value = d.strftime("%m/%d") if d else ""
            ws[f"E{r}"].value = payee
            ws[f"L{r}"].value = desc
            if amount: ws[f"Y{r}"].value = amount
            total_misc += amount

    # 合　計 行24
    _m(ws, "A24:X24");  _label(ws, "A24", "合　　計", size=11, bold=True)
    ws["A24"].border = _b(bot=THIN, left=DBLE, right=THIN)
    _m(ws, "Y24:AC24")
    ws["Y24"].value = total_misc; ws["Y24"].font = _f(bold=True)
    ws["Y24"].alignment = _a("general", "center", wrap=True); _box(ws, 24, 25, 24, 29, bot=THIN)

    # 仮払金（前渡金） 行25
    karibara = data.get("仮払金") or 0
    _m(ws, "A25:X25");  _label(ws, "A25", "仮払金（前渡金）")
    ws["A25"].border = _b(top=THIN, bot=THIN, left=DBLE, right=THIN)
    _m(ws, "Y25:AC25")
    if karibara: ws["Y25"].value = karibara
    ws["Y25"].font = _f(); ws["Y25"].alignment = _a("general", "center"); _box(ws, 25, 25, 25, 29)

    # 差引精算額 行26
    _m(ws, "A26:X26");  _label(ws, "A26", "差引精算額", bold=True)
    ws["A26"].border = _b(bot=DBLE, left=DBLE, right=THIN)
    _m(ws, "Y26:AC26")
    ws["Y26"].value = total_misc - karibara; ws["Y26"].font = _f(bold=True)
    ws["Y26"].alignment = _a("right", "center"); _box(ws, 26, 25, 26, 29, bot=DBLE)

    # （経理課） 行27
    ws.cell(row=27, column=2).value = "（経理課）"
    ws.cell(row=27, column=2).font = _f(); ws.cell(row=27, column=2).alignment = _a("general", "center")

    # 受領日 行27 (右側)
    _m(ws, "R27:T28");  _label(ws, "R27", "受領日")
    ws["R27"].border = _b(top=DBLE)
    _m(ws, "U27:AC28")
    ws["U27"].value = _date.today()   # 受領日は作成日を自動セット
    ws["U27"].number_format = "YYYY/MM/DD"
    ws["U27"].font = _f(); ws["U27"].alignment = _a()
    ws["U27"].border = _b(top=DBLE)

    # 下部印鑑欄 行28-29
    for rng, lbl, c1, c2 in [
        ("B28:F28", "所属長", 2, 6),
        ("G28:K28", "検印",   7, 11),
        ("L28:P28", "出納印", 12, 16),
    ]:
        _m(ws, rng); _label(ws, rng.split(":")[0], lbl); _box(ws, 28, c1, 28, c2)

    _m(ws, "R29:AC30")
    _label(ws, "R29", "上記金額正に領収致しました。")

    for rng in ["B29:F33", "G29:K33", "L29:P33"]:
        c1 = ord(rng[0]) - ord('A') + 1
        c2 = ord(rng.split(":")[1][0]) - ord('A') + 1 + (10 if len(rng.split(":")[1]) > 2 else 0)
        # Use simpler approach:
        pass
    _m(ws, "B29:F33");  _m(ws, "G29:K33");  _m(ws, "L29:P33")
    _box(ws, 29, 2, 33, 6);  _box(ws, 29, 7, 33, 11);  _box(ws, 29, 12, 33, 16)

    # 領収者 行31-33
    _m(ws, "Q31:T33");  _label(ws, "Q31", "領収者\n氏　名", wrap=True)
    ws["Q31"].border = _b(left=THIN)
    _m(ws, "U31:AA33")
    ws["U31"].value = data.get("氏名", ""); ws["U31"].font = _f(11); ws["U31"].alignment = _a()
    _m(ws, "AB31:AC33");  _label(ws, "AB31", "印")

    return wb


# ============================================================
# 仮払申請書
# ============================================================
def build_karibarai(data):
    """
    data keys: 届出日: date, 氏名: str, 目的: str, 金額: int, 使用日: date
    """
    # 金銭受領書と同じ構造をベースに1明細行のみ
    items = []
    if data.get("使用日") and data.get("金額"):
        items = [(data["使用日"], data.get("目的", ""), "", data["金額"])]

    d = {
        "届出日": data.get("届出日"),
        "氏名": data.get("氏名", ""),
        "目的": data.get("目的", ""),
        "items": items,
        "仮払金": 0,
    }
    wb = build_kinsen(d)
    ws = wb.active
    ws.title = "仮払申請書"
    # タイトルを変更
    ws["A1"].value = "仮払申請書"
    return wb
