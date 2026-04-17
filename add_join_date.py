#!/usr/bin/env python3
"""
社員推移スプレッドシートに「入社日」列を追加するスクリプト
月日（G列）の右隣・列H に挿入する

データ源:
  - 三幸プロダクツ: 社員名簿（三幸プロダクツ）.xlsx（正確な入社年月日）
  - 三幸商事:     社員異動記録.xlsx の入社欄より月単位で特定（75期以降）
  - 75期以前入社: データなし（空欄）
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from pathlib import Path
from google_auth import get_sheets_service

SPREADSHEET_ID = '1qZbzvN8goWp1P_ORZuTxeQWnFQblE7B5Xq7k7V4RQY8'
SHEET_ID = 0
BASE_DIR = Path(__file__).parent

# 入社日列の挿入位置（0-based）: 月日(G=6) の直後 → 列H = index 7
INSERT_COL_INDEX = 7

# ===================================================================
# ① 三幸プロダクツ名簿から入社日を自動取得
# ===================================================================
def load_products_join_dates() -> dict:
    """社員名簿（三幸プロダクツ）.xlsx の最新シートから {氏名: 入社日文字列} を返す"""
    path = BASE_DIR / '総務・人事部' / '社員名簿（三幸プロダクツ）.xlsx'
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb['2024.04.01以降']
    result = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = row[3]   # D列: 氏名
        join = row[7]   # H列: 入社年月日
        if not name or not join:
            continue
        name_clean = str(name).replace('\u3000', '').strip()
        if hasattr(join, 'strftime'):
            date_str = join.strftime('%Y/%m/%d')
        else:
            # "2022/10頃" のような文字列もそのまま使う
            date_str = str(join).strip()
        result[name_clean] = date_str
    return result


# ===================================================================
# ② 三幸商事 入社日（社員異動記録から特定済み）
# 75期以前入社のベテラン社員は空欄
# ===================================================================
SANKO_JOIN_DATES = {
    # 75期6月 (2012/06)
    '深井智子':     '2012/06',

    # 77期9月 (2014/09)
    '片岡源太郎':   '2014/09',

    # 78期10月 (2015/10)
    '秋月雅一':     '2015/10',

    # 80期10月 (2017/10)
    '柴田航':       '2017/10',

    # 81期7月 (2018/07)  ※有期→正社員は2018/10
    '大浦壮司':     '2018/07',

    # 81期4月 (2019/04/01)
    '倉元太郎':     '2019/04/01',
    '竹原黎':       '2019/04/01',
    '菅頌子':       '2019/04/01',   # 婚姻後→加茂頌子
    '船吉紗矢':     '2019/04/01',

    # 81期5月 (2019/05)
    '福田正志':     '2019/05/13',
    '田中マユミ':   '2019/05/20',

    # 82期4月 (2020/04/01)
    '池ヶ谷武弘':   '2020/04/01',
    '飯塚理彩':     '2020/04/01',

    # 84期5月 (2021/05/10)
    '鈴木祐介':     '2021/05/10',

    # 85期4月 (2022/04/01)
    '大﨑紗佳':     '2022/04/01',
    '片岡隼人':     '2022/04/01',

    # 86期4月 (2023/04/01)
    '角野百香':     '2023/04/01',
}


def main():
    # プロダクツ名簿から入社日取得
    products_dates = load_products_join_dates()
    print(f'三幸プロダクツ社員 入社日取得: {len(products_dates)}名')

    # 全入社日マップを統合（プロダクツ優先、三幸商事はSANKO_JOIN_DATESから）
    all_dates: dict = {}
    all_dates.update(SANKO_JOIN_DATES)
    all_dates.update(products_dates)   # プロダクツは氏名スペースなしで格納済み

    service = get_sheets_service()

    # ---------- 1. 現在の氏名列を取得 ----------
    result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range='シート1!A:A'
    ).execute()
    name_rows = result.get('values', [])
    print(f'スプレッドシート行数: {len(name_rows)}')

    # ---------- 2. 列H を挿入（G=月日 の直後） ----------
    service.spreadsheets().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={
            'requests': [{
                'insertDimension': {
                    'range': {
                        'sheetId': SHEET_ID,
                        'dimension': 'COLUMNS',
                        'startIndex': INSERT_COL_INDEX,
                        'endIndex': INSERT_COL_INDEX + 1,
                    },
                    'inheritFromBefore': True
                }
            }]
        }
    ).execute()
    print('列H を挿入しました（G=月日 の右隣）')

    # ---------- 3. ヘッダーとデータを書き込む ----------
    data_updates = [
        {'range': 'シート1!H1', 'values': [['']]},
        {'range': 'シート1!H2', 'values': [['']]},
        {'range': 'シート1!H3', 'values': [['入社日']]},
    ]

    written = 0
    blanks = []
    for i, row in enumerate(name_rows):
        row_num = i + 1
        if row_num <= 3:
            continue
        if not row or not row[0].strip():
            continue
        name = row[0].strip()
        # 氏名の全角スペースを除去して照合
        name_key = name.replace('\u3000', '').strip()

        date_str = all_dates.get(name) or all_dates.get(name_key) or ''
        if date_str:
            data_updates.append({
                'range': f'シート1!H{row_num}',
                'values': [[date_str]]
            })
            written += 1
            print(f'  ✓ {name}: {date_str}')
        else:
            blanks.append(name)

    # バッチ書き込み
    service.spreadsheets().values().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={
            'valueInputOption': 'USER_ENTERED',
            'data': data_updates
        }
    ).execute()

    print(f'\n=== 完了 ===')
    print(f'入社日を記入した社員: {written} 名')
    print(f'入社日 空欄（データなし）: {len(blanks)} 名')
    if blanks:
        print('  空欄:', ', '.join(blanks))


if __name__ == '__main__':
    main()
